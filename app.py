from flask import Flask, session, render_template, request, redirect, url_for
from flask_session import Session
from flask import flash
from flask import Flask
from flask import request
from PIL import Image
from io import BytesIO
import models
from sqlalchemy import and_
from flask import flash
from flask import Flask
from flask import request
from flask import redirect
from flask import url_for
from flask import render_template
from database import engine
from database import Database
from database import db_session
import uuid
import models
import os
from datetime import datetime
import openpyxl
from flask import Flask, render_template, request, make_response
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from math import ceil
from flask import Flask, send_file, make_response
from io import BytesIO
import sqlite3
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from flask import Flask, send_file
import shutil
from flask import Flask, g
import sqlite3

engine = create_engine('sqlite:///bd_beneficiarios.db')


SECRET_KEY = 'cvbnvcsdfgbng45623890'

#carpetas imagenes
FOTOS_USUARIOS = 'static/images/usuarios/'
ALLOWED_IMAGE_TYPES = ['png', 'jpg', 'jpeg', 'gif']

app= Flask(__name__)
app.config.from_object(__name__)
Database.metadata.create_all(engine)

app.config['DATABASE'] = 'sqlite:///bd_beneficiarios.db'

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

#proceso validación imagenes
SESSION_TYPE = 'filesystem'
app.config.from_object(__name__)
Session(app)

def validate_file_type(filename: str, types):
    return filename != '' and \
        '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in types

def save_game_image(file, game_path):
    image_data = file.read()
    image = Image.open(BytesIO(image_data))

    file_name = "%s.png" % str(uuid.uuid4())
    file_path = game_path + file_name

    image.save(file_path, format="png")

    return file_name

#administrador datos sesión
admin_id = 'Administrador'
admin_passw = 'toor#45'

@app.get('/set/')
def set():
    session['sesion_iniciada'] = 'Mitzi'
    return 'ok'

#sesiones
def get_session_type():
    return session.get('type', None)

def get_user_from_session():
    user_id = session.get('user_id', None)

    if user_id:
        return db_session.query(models.Usuarios).get(user_id)
    
    return None

@app.teardown_request
def remove_session(ex=None):
    db_session.remove()

#validar sesión
@app.get('/sesion')
def sesion():
    if 'nombre_usuario' in session:
        nombre_completo = usuario.nombre  # Obtén el nombre completo del usuario
        nombre = nombre_completo.split()  # Divide el nombre en palabras
        foto_usuario = session['usuario_foto']
        foto = foto_usuario.split()[0]
        return render_template('sesion.html', nombre_completo=nombre)
    else:
        return redirect(url_for('login_form'))

#Destruir sesión
@app.get('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_form'))

@app.get('/')   
def login_form():
    return render_template('sesion.html')

@app.post('/login')
def login_post():
    usuario = request.form['usuario']
    password = request.form['password']

    # Validar usuario y contraseña de administrador
    if usuario == admin_id and password == admin_passw:
        session['user_type'] = 'admin'
        session['nombre_usuario'] = 'Administrador'
        return redirect(url_for('principal'))  # Redirigir al panel de administrador

    # Validar usuario y contraseña de usuario normal
    usuario = db_session.query(models.Usuario).filter(
        and_(models.Usuario.usuario == usuario, models.Usuario.contrasena == password)
    ).first()

    if usuario is None:
        flash('Credenciales inválidas. Inténtalo de nuevo.', 'error')
        return redirect(url_for('login_form'))

    session['user_type'] = 'normal'
    session['user_id'] = usuario.id
    session['nombre_completo'] = usuario.nombre
    session['foto_usuario'] = usuario.foto # Establecer el nombre de usuario en la sesión para el usuario normal
    session['area'] = usuario.area  # Almacena el área del usuario en la sesión

    # Determina a qué módulos tiene acceso el usuario en función del área seleccionada
    modulos = []

    if session['area'] == 'Desayuno escolar':
        modulos = ['Desayunos fríos', 'Desayunos calientes']
        session['modulos'] = modulos  # Almacena los módulos a los que el usuario tiene acceso en la sesión
        return redirect(url_for('desayunos')) 
    elif session['area'] == 'Horta DIF':
        modulos = ['Huertos escolares', 'Huertos familiares']
    

    session['modulos'] = modulos  # Almacena los módulos a los que el usuario tiene acceso en la sesión

    return redirect(url_for('huertos'))  # Redirige al panel principal

@app.route('/respaldo_bd')
def respaldo_basedatos():
    nombre_archivo_original = 'bd_beneficiarios.db'
    nombre_archivo_respaldo = 'respaldo_basedatos.sqlite'

    # Copiar el archivo de la base de datos original al de respaldo
    try:
        shutil.copy(nombre_archivo_original, nombre_archivo_respaldo)
        return send_file(nombre_archivo_respaldo, as_attachment=True)
    except Exception as e:
        return str(e)
    
# Configuración modulo principal del sistema
@app.get('/principal')
def principal():
    # Verificar si el usuario ha iniciado sesión como administrador o usuario normal
    if 'user_type' in session:
        user_type = session['user_type']
        if user_type == 'admin':
            # El administrador tiene acceso a la ruta principal
            return render_template('home.html')
    
    # Si no se ha iniciado sesión o no se es administrador, redirigir a la página de inicio de sesión
    return redirect(url_for('login_form'))

# Configuración modulo principal del sistema
@app.get('/Desayunos/escolares')
def desayunos():
    # Verificar si el usuario ha iniciado sesión como administrador o usuario normal
    if 'user_type' in session:
        user_type = session['user_type']
        if user_type == 'normal':
            # El administrador tiene acceso a la ruta principal
            return render_template('desayunos.html')
    
    # Si no se ha iniciado sesión o no se es administrador, redirigir a la página de inicio de sesión
    return redirect(url_for('login_form'))

# Configuración modulo principal del sistema
@app.get('/Horta/DIF')
def huertos():
    # Verificar si el usuario ha iniciado sesión como administrador o usuario normal
    if 'user_type' in session:
        user_type = session['user_type']
        if user_type == 'normal':
            # El administrador tiene acceso a la ruta principal
            return render_template('huertos.html')
    
    # Si no se ha iniciado sesión o no se es administrador, redirigir a la página de inicio de sesión
    return redirect(url_for('login_form'))

# Configuración formulario actualizar perfil
@app.route('/perfil')
def perfil():
    if 'user_id' in session:  # Verificar que el usuario esté en sesión
        user_id = session['user_id']
        usuario = db_session.query(models.Usuario).get(user_id)  # Obtener el usuario actual
        if usuario:  # Verificar que se haya encontrado el usuario
            user_type = session.get('user_type', 'normal')  # Establecer un valor predeterminado si 'user_type' no está en sesión
            if user_type == 'normal':
                return render_template('formulario-actualizar-perfil.html', usuario=usuario)
    
    return redirect(url_for('login_form'))

# Configuración formulario actualizar perfil
@app.route('/datos')
def datos():
    if 'user_id' in session:  # Verificar que el usuario esté en sesión
        user_id = session['user_id']
        usuario = db_session.query(models.Usuario).get(user_id)  # Obtener el usuario actual
        if usuario:  # Verificar que se haya encontrado el usuario
            user_type = session.get('user_type', 'normal')  # Establecer un valor predeterminado si 'user_type' no está en sesión
            if user_type == 'normal':
                return render_template('formulario-actualizar-perfil-datos.html', usuario=usuario)
    
    return redirect(url_for('login_form'))

# Actualización
@app.route('/perfil/<int:id>/update', methods=['POST'])
def perfil_update(id):
    if 'user_id' in session:
        user_id = session['user_id']
        if id == user_id:
            Usuario = db_session.query(models.Usuario).get(id)
            nuevo_nombre_usuario = request.form['nuevo-nombre-usuario']
            nuevos_apellidos = request.form['nuevos-apellidos']
            nuevo_usuario = request.form['nuevo-usuario']
            nueva_imagen_usuario = request.files['nueva-imagen-usuario']

            if nuevo_nombre_usuario:
                Usuario.nombre = nuevo_nombre_usuario
                session['nombre_completo'] = nuevo_nombre_usuario  # Actualiza el nombre en la sesión
            if nuevos_apellidos:
                Usuario.apellidos = nuevos_apellidos
            if nuevo_usuario:
                Usuario.usuario = nuevo_usuario
            if nueva_imagen_usuario:
                # Guardar la nueva imagen en el servidor
                nueva_imagen_usuario.save(os.path.join(app.root_path, 'static/images/usuarios/', nueva_imagen_usuario.filename))
                Usuario.foto = nueva_imagen_usuario.filename
            db_session.add(Usuario)
            db_session.commit()

    return redirect(url_for('desayunos'))

# Actualización
@app.route('/usuario/perfil/<int:id>/update', methods=['POST'])
def datos_update(id):
    if 'user_id' in session:
        user_id = session['user_id']
        if id == user_id:
            Usuario = db_session.query(models.Usuario).get(id)
            nuevo_nombre_usuario = request.form['nuevo-nombre-usuario']
            nuevos_apellidos = request.form['nuevos-apellidos']
            nuevo_usuario = request.form['nuevo-usuario']
            nueva_imagen_usuario = request.files['nueva-imagen-usuario']

            if nuevo_nombre_usuario:
                Usuario.nombre = nuevo_nombre_usuario
                session['nombre_completo'] = nuevo_nombre_usuario  # Actualiza el nombre en la sesión
            if nuevos_apellidos:
                Usuario.apellidos = nuevos_apellidos
            if nuevo_usuario:
                Usuario.usuario = nuevo_usuario
            if nueva_imagen_usuario:
                # Guardar la nueva imagen en el servidor
                nueva_imagen_usuario.save(os.path.join(app.root_path, 'static/images/usuarios/', nueva_imagen_usuario.filename))
                Usuario.foto = nueva_imagen_usuario.filename
            db_session.add(Usuario)
            db_session.commit()

    return redirect(url_for('huertos'))
@app.get('/registrar/usuario')
def usuario():
    # Verificar si el usuario ha iniciado sesión como administrador o usuario normal
    if 'user_type' in session:
        user_type = session['user_type']
        if user_type == 'admin':
            return render_template('formulario-registrar-usuario.html')
    return redirect(url_for('login_form'))

from sqlalchemy import func

@app.get('/usuarios')
def usuarios():
    # Verificar si el usuario ha iniciado sesión como administrador o usuario normal
    if 'user_type' in session:
        user_type = session['user_type']
        if user_type == 'admin':
            # Consulta para contar el número total de registros de usuarios
            total_registros = db_session.query(func.count(models.Usuario.id)).scalar()

            # Obtener el parámetro 'per_page' de la solicitud GET
            per_page = request.args.get('per_page', type=int, default=5)
            page = request.args.get('page', type=int, default=1)

            start_index = (page - 1) * per_page
            end_index = start_index + per_page

            # Obtener usuarios paginados
            usuario_paginados = db_session.query(models.Usuario).slice(start_index, end_index)

            total_pages = total_registros // per_page
            if total_registros % per_page != 0:
                total_pages += 1

            if page < 1:
                page = 1
            elif page > total_pages:
                page = total_pages

            return render_template('modulo-usuarios-administrador.html', Usuario=usuario_paginados, page=page, total_pages=total_pages, per_page=per_page, total_registros=total_registros)

    return redirect(url_for('login_form'))

@app.get('/usuarios/form')
def usuarios_form():
    usuario = db_session.query(models.Usuario).all()
    if 'user_type' in session:
        user_type = session['user_type']
        if user_type == 'admin':
            return render_template('formulario-registrar-usuario.html', Usuario = usuario)
    return redirect(url_for('login_form'))
#post
@app.post('/usuarios/post')
def usuarios_post():
    nombre = request.form['nombre']
    apellidos = request.form['apellidos']
    usuario = request.form['usuario']
    contrasena = request.form['password']
    area = request.form['area']
    foto = 'default.png'

    if 'imagen-usuario' not in request.files:
         flash('Imagen no cargada')
         return redirect(url_for('usuarios'))

    image_cover_file = request.files['imagen-usuario']

    if not validate_file_type(image_cover_file.filename, ALLOWED_IMAGE_TYPES):
        flash('Tipo de extension invalida')
        return redirect(url_for('usuarios'))

    if nombre == None or nombre =='':       
        flash("No se registro el usuario") 
        return redirect(url_for('usuarios'))

    foto = save_game_image(image_cover_file, FOTOS_USUARIOS)


    registro_nuevo =models.Usuario(
        nombre = nombre,
        apellidos = apellidos,
        usuario = usuario,
        contrasena = contrasena ,
        area = area,
        foto = foto)
    
    db_session.add(registro_nuevo)
    db_session.commit()
    return redirect(url_for('usuarios'))

#Eliminar
@app.get('/usuarios/<id>/delete')
def usuarios_delete(id):
    usuario = db_session.query(models.Usuario).get(id)
    db_session.delete(usuario)
    db_session.commit()
    return redirect(url_for('usuarios'))

#actualización
@app.post('/usuarios/<id>/update')
def usuarios_update(id):
    Usuario = db_session.query(models.Usuario).get(id)
    nuevo_nombre_usuario = request.form['nuevo-nombre-usuario']
    nuevos_apellidos = request.form['nuevos-apellidos']
    nuevo_usuario = request.form['nuevo-usuario']
    nueva_area = request.form['nueva-area']
    nueva_imagen_usuario = request.files['nueva-imagen-usuario']
    
    if nuevo_nombre_usuario != None and nuevo_nombre_usuario != "":
        Usuario.nombre = nuevo_nombre_usuario
    if nuevos_apellidos != None and nuevos_apellidos != "":
        Usuario.apellidos = nuevos_apellidos
    if nuevo_usuario != None and nuevo_usuario != "":
        Usuario.usuario = nuevo_usuario
    if  nueva_area != None and  nueva_area != "":
        Usuario.area =  nueva_area
    
    if nueva_imagen_usuario:
    # Guardar la nueva imagen en el servidor
        nueva_imagen_usuario.save(os.path.join(app.root_path, 'static/images/usuarios/', nueva_imagen_usuario.filename))
        Usuario.foto = nueva_imagen_usuario.filename
    nueva_imagen_usuario = request.files.get('nueva-imagen-usuario')
    
    db_session.add(Usuario)
    db_session.commit()

    return redirect(url_for('usuarios'))

@app.route('/desayunos_frios')
def frios():
    frio = db_session.query(models.Frio).all()
    user_type = session.get('user_type')
    
    # Obtener el parámetro 'per_page' de la solicitud GET
    per_page = request.args.get('per_page', type=int, default=5)
    page = request.args.get('page', type=int, default=1)
    
    total_records = db_session.query(func.count(models.Frio.id)).scalar()  # Contar el número total de registros

    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    total_pages = total_records // per_page
    if total_records % per_page != 0:
        total_pages += 1

    if page < 1:
        page = 1
    elif page > total_pages:
        page = total_pages
    
    if user_type == 'admin':
        # Los administradores pueden acceder a todas las áreas y módulos
        frio_paginados = db_session.query(models.Frio).slice(start_index, end_index)
        return render_template('modulo-desayuno-frio.html', Frio=frio_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)

    # Si el usuario es normal, verifica si tiene acceso al módulo específico
    modulos_accesibles = session.get('modulos', [])
    if 'Desayunos fríos' in modulos_accesibles:
        # Los usuarios normales también pueden paginar los resultados
        frio_paginados = db_session.query(models.Frio).slice(start_index, end_index)
        return render_template('modulo-desayuno-frio.html', Frio=frio_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)
    else:
        return redirect(url_for('login_form'))

#formulario desayunos frios
@app.get('/desayunos/frios/form')
def frios_form():
    user_type = session.get('user_type')
    area = session.get('area')
    modulos = session.get('modulos', [])

    if user_type == 'admin':
        frio = db_session.query(models.Frio).all()
        return render_template('formulario-desayunos-frios.html', Frio=frio)

    if user_type == 'normal' and area == 'Desayuno escolar' and 'Desayunos fríos' in modulos:
        frio = db_session.query(models.Frio).all()
        return render_template('formulario-desayunos-frios.html', Frio=frio)
    
    return redirect(url_for('login_form'))

#post
@app.post('/desayunos/frios/post')
def frios_post():
    nombre_escuela = request.form['nombre_escuela']
    cct = request.form['cct']
    localidad_escuela = request.form['localidad_escuela']
    nombre_beneficiario = request.form['nombre_beneficiario']
    escolaridad_beneficiario = request.form['escolaridad_beneficiario']
    talla_beneficiario = request.form['talla_beneficiario']
    curp_beneficiario = request.form['curp_beneficiario']
    peso_beneficiario = request.form['peso_beneficiario']
    referencias_domicilio = request.form['referencias_domicilio']
    nombre_padre = request.form['nombre_padre']
    escolaridad_padre = request.form['escolaridad_padre']
    ine_padre = request.form['ine_padre']
    curp_padre = request.form['curp_padre']
    telefono_padre = request.form['telefono_padre']
    estado_civil = request.form['estado_civil']

    registro_nuevo =models.Frio(
        nombre_escuela = nombre_escuela,
        cct = cct,
        localidad_escuela = localidad_escuela,
        nombre_beneficiario = nombre_beneficiario ,
        escolaridad_beneficiario = escolaridad_beneficiario,
        talla_beneficiario = talla_beneficiario,
        curp_beneficiario = curp_beneficiario,
        peso_beneficiario = peso_beneficiario,
        referencias_domicilio = referencias_domicilio,
        nombre_padre = nombre_padre,
        escolaridad_padre = escolaridad_padre,
        ine_padre = ine_padre,
        curp_padre = curp_padre,
        telefono_padre = telefono_padre,
        estado_civil = estado_civil,
    )
    
    db_session.add(registro_nuevo)
    db_session.commit()
    return redirect(url_for('frios'))

#Eliminar
@app.get('/desayunos/frios/<id>/delete')
def frio_delete(id):
    frio = db_session.query(models.Frio).get(id)
    db_session.delete(frio)
    db_session.commit()
    return redirect(url_for('frios'))

#actualización
@app.post('/desayunos/frios/<id>/update')
def frio_update(id):
    Frio = db_session.query(models.Frio).get(id)
    nuevo_nombre_escuela = request.form['nuevo-nombre-escuela']
    nuevo_cct = request.form['nuevo-cct']
    nueva_localidad_escuela = request.form['nueva-localidad-escuela']
    nuevo_nombre_beneficiario = request.form['nuevo-nombre-beneficiario']
    nueva_escolaridad_beneficiario= request.form['nueva-escolaridad-beneficiario']
    nueva_talla_beneficiario = request.form['nueva-talla-beneficiario']
    nueva_curp_beneficiario = request.form['nueva-curp-beneficiario']
    nuevo_peso_beneficiario = request.form['nuevo-peso-beneficiario']
    nueva_referencia_domicilio = request.form['nueva-referencia-domicilio']
    nuevo_nombre_padre = request.form['nuevo-nombre-padre']
    nueva_escolaridad_padre = request.form['nueva-escolaridad-padre']
    nuevo_ine_padre = request.form['nuevo-ine-padre']
    nuevo_curp_padre = request.form['nuevo-curp-padre']
    nuevo_telefono_padre = request.form['nuevo-telefono-padre']
    nuevo_estado_civil = request.form['nuevo-estado-civil']
    
    if nuevo_nombre_escuela != None and nuevo_nombre_escuela != "":
        Frio.nombre_escuela = nuevo_nombre_escuela
    if nuevo_cct != None and nuevo_cct != "":
        Frio.cct = nuevo_cct
    if nueva_localidad_escuela != None and nueva_localidad_escuela != "":
        Frio.localidad_escuela = nueva_localidad_escuela
    if nuevo_nombre_beneficiario != None and nuevo_nombre_beneficiario != "":
        Frio.nombre_beneficiario = nuevo_nombre_beneficiario
    if nueva_escolaridad_beneficiario != None and nueva_escolaridad_beneficiario != "":
        Frio.escolaridad_beneficiario = nueva_escolaridad_beneficiario
    if nueva_talla_beneficiario != None and nueva_talla_beneficiario != "":
        Frio.talla_beneficiario = nueva_talla_beneficiario
    if nueva_curp_beneficiario != None and nueva_curp_beneficiario != "":
        Frio.curp_beneficiario = nueva_curp_beneficiario
    if nuevo_peso_beneficiario != None and nuevo_peso_beneficiario != "":
        Frio.peso_beneficiario = nuevo_peso_beneficiario
    if nueva_referencia_domicilio != None and nueva_referencia_domicilio != "":
        Frio.referencias_domicilio = nueva_referencia_domicilio
    if nuevo_nombre_padre != None and nuevo_nombre_padre != "":
        Frio.nombre_padre = nuevo_nombre_padre
    if nueva_escolaridad_padre != None and nueva_escolaridad_padre != "":
        Frio.escolaridad_padre = nueva_escolaridad_padre
    if nuevo_ine_padre != None and nuevo_ine_padre != "":
        Frio.ine_padre = nuevo_ine_padre
    if nuevo_curp_padre != None and nuevo_curp_padre != "":
        Frio.curp_padre = nuevo_curp_padre
    if nuevo_telefono_padre != None and nuevo_telefono_padre != "":
        Frio.telefono_padre = nuevo_telefono_padre
    if nuevo_estado_civil != None and nuevo_estado_civil != "":
        Frio.estado_civil = nuevo_estado_civil
    
    db_session.add(Frio)
    db_session.commit()

    return redirect(url_for('frios'))

@app.get('/desayunos/escolares/comunitarios')
def comunitario():
    comunitario = db_session.query(models.Comunitario).all()
    user_type = session.get('user_type')
    
    # Obtener el parámetro 'per_page' de la solicitud GET
    per_page = request.args.get('per_page', type=int, default=5)
    page = request.args.get('page', type=int, default=1)
    
    total_records = db_session.query(func.count(models.Comunitario.id)).scalar()  # Contar el número total de registros

    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    total_pages = total_records // per_page
    if total_records % per_page != 0:
        total_pages += 1

    if page < 1:
        page = 1
    elif page > total_pages:
        page = total_pages
    
    if user_type == 'admin':
        # Los administradores pueden acceder a todas las áreas y módulos
        comunitario_paginados = db_session.query(models.Comunitario).slice(start_index, end_index)
        return render_template('modulo-desayuno-caliente.html', Comunitario=comunitario_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)

    modulos_accesibles = session.get('modulos', [])
    if 'Desayunos calientes' in modulos_accesibles:
        # Los usuarios normales también pueden paginar los resultados
        comunitario_paginados = db_session.query(models.Comunitario).slice(start_index, end_index)
        return render_template('modulo-desayuno-caliente.html', Comunitario=comunitario_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)
    else:
        return redirect(url_for('login_form'))
    
#formulario desayunos calientes
@app.get('/desayunos/escolares/comunitarios/form')
def comunitario_form():
    comunitario = db_session.query(models.Comunitario).all()
    user_type = session.get('user_type')

    if user_type == 'admin':
        return render_template('formulario-desayunos-calientes.html', Comunitario=comunitario)

    modulos_accesibles = session.get('modulos', [])
    if 'Desayunos calientes' in modulos_accesibles:
        return render_template('formulario-desayunos-calientes.html', Comunitario=comunitario)
    else:
        return redirect(url_for('login_form'))
#post
@app.post('/desayunos/escolares/comunitarios/post')
def comunitario_post():
    nombre_escuela = request.form['nombre_escuela']
    cct = request.form['cct']
    localidad_escuela = request.form['localidad_escuela']
    nombre_beneficiario = request.form['nombre_beneficiario']
    escolaridad_beneficiario = request.form['escolaridad_beneficiario']
    talla_beneficiario = request.form['talla_beneficiario']
    curp_beneficiario = request.form['curp_beneficiario']
    peso_beneficiario = request.form['peso_beneficiario']
    referencias_domicilio = request.form['referencias_domicilio']
    nombre_padre = request.form['nombre_padre']
    escolaridad_padre = request.form['escolaridad_padre']
    ine_padre = request.form['ine_padre']
    curp_padre = request.form['curp_padre']
    telefono_padre = request.form['telefono_padre']
    estado_civil = request.form['estado_civil']

    registro_nuevo =models.Comunitario(
        nombre_escuela = nombre_escuela,
        cct = cct,
        localidad_escuela = localidad_escuela,
        nombre_beneficiario = nombre_beneficiario ,
        escolaridad_beneficiario = escolaridad_beneficiario,
        talla_beneficiario = talla_beneficiario,
        curp_beneficiario = curp_beneficiario,
        peso_beneficiario = peso_beneficiario,
        referencias_domicilio = referencias_domicilio,
        nombre_padre = nombre_padre,
        escolaridad_padre = escolaridad_padre,
        ine_padre = ine_padre,
        curp_padre = curp_padre,
        telefono_padre = telefono_padre,
        estado_civil = estado_civil,
    )
    
    db_session.add(registro_nuevo)
    db_session.commit()
    return redirect(url_for('comunitario'))

#Eliminar
@app.get('/desayunos/escolares/comunitarios/<id>/delete')
def comunitario_delete(id):
    comunitario = db_session.query(models.Comunitario).get(id)
    db_session.delete(comunitario)
    db_session.commit()
    return redirect(url_for('comunitario'))

#actualización
@app.post('/desayunos/escolares/comunitarios/<id>/update')
def comunitario_update(id):
    Comunitario = db_session.query(models.Comunitario).get(id)
    nuevo_nombre_escuela = request.form['nuevo-nombre-escuela']
    nuevo_cct = request.form['nuevo-cct']
    nueva_localidad_escuela = request.form['nueva-localidad-escuela']
    nuevo_nombre_beneficiario = request.form['nuevo-nombre-beneficiario']
    nueva_escolaridad_beneficiario= request.form['nueva-escolaridad-beneficiario']
    nueva_talla_beneficiario = request.form['nueva-talla-beneficiario']
    nueva_curp_beneficiario = request.form['nueva-curp-beneficiario']
    nuevo_peso_beneficiario = request.form['nuevo-peso-beneficiario']
    nueva_referencia_domicilio = request.form['nueva-referencia-domicilio']
    nuevo_nombre_padre = request.form['nuevo-nombre-padre']
    nueva_escolaridad_padre = request.form['nueva-escolaridad-padre']
    nuevo_ine_padre = request.form['nuevo-ine-padre']
    nuevo_curp_padre = request.form['nuevo-curp-padre']
    nuevo_telefono_padre = request.form['nuevo-telefono-padre']
    nuevo_estado_civil = request.form['nuevo-estado-civil']
    
    if nuevo_nombre_escuela != None and nuevo_nombre_escuela != "":
        Comunitario.nombre_escuela = nuevo_nombre_escuela
    if nuevo_cct != None and nuevo_cct != "":
        Comunitario.cct = nuevo_cct
    if nueva_localidad_escuela != None and nueva_localidad_escuela != "":
        Comunitario.localidad_escuela = nueva_localidad_escuela
    if nuevo_nombre_beneficiario != None and nuevo_nombre_beneficiario != "":
        Comunitario.nombre_beneficiario = nuevo_nombre_beneficiario
    if nueva_escolaridad_beneficiario != None and nueva_escolaridad_beneficiario != "":
        Comunitario.escolaridad_beneficiario = nueva_escolaridad_beneficiario
    if nueva_talla_beneficiario != None and nueva_talla_beneficiario != "":
        Comunitario.talla_beneficiario = nueva_talla_beneficiario
    if nueva_curp_beneficiario != None and nueva_curp_beneficiario != "":
        Comunitario.curp_beneficiario = nueva_curp_beneficiario
    if nuevo_peso_beneficiario != None and nuevo_peso_beneficiario != "":
        Comunitario.peso_beneficiario = nuevo_peso_beneficiario
    if nueva_referencia_domicilio != None and nueva_referencia_domicilio != "":
        Comunitario.referencias_domicilio = nueva_referencia_domicilio
    if nuevo_nombre_padre != None and nuevo_nombre_padre != "":
        Comunitario.nombre_padre = nuevo_nombre_padre
    if nueva_escolaridad_padre != None and nueva_escolaridad_padre != "":
        Comunitario.escolaridad_padre = nueva_escolaridad_padre
    if nuevo_ine_padre != None and nuevo_ine_padre != "":
        Comunitario.ine_padre = nuevo_ine_padre
    if nuevo_curp_padre != None and nuevo_curp_padre != "":
        Comunitario.curp_padre = nuevo_curp_padre
    if nuevo_telefono_padre != None and nuevo_telefono_padre != "":
        Comunitario.telefono_padre = nuevo_telefono_padre
    if nuevo_estado_civil != None and nuevo_estado_civil != "":
        Comunitario.estado_civil = nuevo_estado_civil
    
    db_session.add(Comunitario)
    db_session.commit()

    return redirect(url_for('comunitario'))

@app.get('/huertos_escolares')
def escolares():
    escolares = db_session.query(models.Escolares).all()
    user_type = session.get('user_type')
    
    # Obtener el parámetro 'per_page' de la solicitud GET
    per_page = request.args.get('per_page', type=int, default=5)
    page = request.args.get('page', type=int, default=1)
    
    total_records = db_session.query(func.count(models.Escolares.id_beneficiario)).scalar()  # Contar el número total de registros

    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    total_pages = total_records // per_page
    if total_records % per_page != 0:
        total_pages += 1

    if page < 1:
        page = 1
    elif page > total_pages:
        page = total_pages
    
    if user_type == 'admin':
        # Los administradores pueden acceder a todas las áreas y módulos
        escolares_paginados = db_session.query(models.Escolares).slice(start_index, end_index)
        return render_template('modulo-huerto-escolar.html', Escolares=escolares_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)

    modulos_accesibles = session.get('modulos', [])
    if 'Huertos escolares' in modulos_accesibles:
        # Los usuarios normales también pueden paginar los resultados
        escolares_paginados = db_session.query(models.Escolares).slice(start_index, end_index)
        return render_template('modulo-huerto-escolar.html', Escolares=escolares_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)
    else:
        return redirect(url_for('login_form'))

#formulario huertos escolares
@app.get('/huertos/escolares/form')
def escolares_form():
    escolares = db_session.query(models.Escolares).all()
    user_type = session.get('user_type')

    if user_type == 'admin':
        return render_template('formulario-huertos-escolares.html', Escolares=escolares)

    modulos_accesibles = session.get('modulos', [])
    if 'Huertos escolares' in modulos_accesibles:
        return render_template('formulario-huertos-escolares.html', Escolares=escolares)
    else:
        return redirect(url_for('login_form'))

#post
@app.post('/huertos/escolares/post')
def escolares_post():
    nombre_escuela = request.form['nombre_escuela']
    cct = request.form['cct']
    localidad_escuela = request.form['localidad_escuela']
    nombre_beneficiario = request.form['nombre_beneficiario']
    fecha_str = request.form['fecha_nacimiento']
    fecha_otj = datetime.strptime(fecha_str, '%Y-%m-%d')
    genero = request.form['genero']
    estado_civil = request.form['estado_civil']
    escolaridad_beneficiario = request.form['escolaridad_beneficiario']
    ine_beneficiario = request.form['ine_beneficiario']
    curp_beneficiario = request.form['curp_beneficiario']
    numero_telefono = request.form['numero_telefono']
    correo_beneficiario = request.form['correo_beneficiario']
    municipio_beneficiario = request.form['municipio_beneficiario']
    localidad_beneficiario = request.form['localidad_beneficiario']
    colonia_beneficiario = request.form['colonia_beneficiario']
    codigo_postal = request.form['codigo_postal']
    numero_exterior = request.form['numero_exterior']
    numero_interior = request.form['numero_interior']
    calle = request.form['calle']
    referencias_domicilio = request.form['referencias_domicilio']
    tipo_apoyo = request.form['tipo_apoyo']
    criterio_seleccion = request.form['criterio_seleccion']
    periodo_apoyo = request.form['periodo_apoyo']
    fecha_asignada = request.form['fecha_registro']
    fecha_obj = datetime.strptime(fecha_asignada, '%Y-%m-%d')
    
    registro_nuevo =models.Escolares(
        nombre_escuela = nombre_escuela,
        cct = cct,
        localidad_escuela = localidad_escuela,
        nombre_beneficiario = nombre_beneficiario,
        fecha_nacimiento = fecha_otj,
        genero = genero,
        estado_civil = estado_civil,
        escolaridad_beneficiario = escolaridad_beneficiario,
        ine_beneficiario = ine_beneficiario,
        curp_beneficiario = curp_beneficiario,
        numero_telefono =numero_telefono,
        correo_beneficiario = correo_beneficiario,
        municipio_beneficiario = municipio_beneficiario,
        localidad_beneficiario = localidad_beneficiario,
        colonia_beneficiario = colonia_beneficiario,
        codigo_postal = codigo_postal,
        numero_exterior =  numero_exterior,
        numero_interior = numero_interior,
        calle = calle,
        referencias_domicilio = referencias_domicilio,
        tipo_apoyo = tipo_apoyo,
        criterio_seleccion = criterio_seleccion,
        periodo_apoyo = periodo_apoyo,
        fecha_registro = fecha_obj
        
    )
    
    db_session.add(registro_nuevo)
    db_session.commit()
    return redirect(url_for('escolares'))

#Eliminar
@app.get('/huertos/escolares/<id>/delete')
def escolares_delete(id):
    escolares = db_session.query(models.Escolares).get(id)
    db_session.delete(escolares)
    db_session.commit()
    return redirect(url_for('escolares'))

#actualización
@app.post('/huertos/escolares/<id>/update')
def escolares_update(id):
    Escolares = db_session.query(models.Escolares).get(id)
    nuevo_nombre_escuela = request.form['nuevo-nombre-escuela']
    nuevo_cct = request.form['nuevo-cct']
    nueva_localidad_escuela = request.form['nueva-localidad-escuela']
    nuevo_nombre_beneficiario = request.form['nuevo-nombre-beneficiario']
    nueva_fecha_str = request.form['nueva-fecha-nacimiento']
    nueva_fecha_otj = datetime.strptime(nueva_fecha_str, '%Y-%m-%d')
    nuevo_genero_beneficiario = request.form['nuevo-genero-beneficiario']
    nuevo_estado_civil = request.form['nuevo-estado-civil']
    nueva_escolaridad_beneficiario = request.form['nueva-escolaridad-beneficiario']
    nuevo_ine_beneficiario = request.form['nuevo-ine-beneficiario']
    nueva_curp_beneficiario = request.form['nueva-curp-beneficiario']
    nuevo_telefono = request.form['nuevo-telefono']
    nuevo_correo = request.form['nuevo-correo']
    nuevo_municipio = request.form['nuevo-municipio']
    nueva_localidad = request.form['nueva-localidad']
    nueva_colonia = request.form['nueva-colonia']
    nuevo_codigo_postal = request.form['nuevo-codigo-postal']
    nuevo_numero_exterior = request.form['nuevo-numero-exterior']
    nuevo_numero_interior = request.form['nuevo-numero-interior']
    nueva_calle = request.form['nueva-calle']
    nueva_referencia_domicilio = request.form['nueva-referencia-domicilio']
    nuevo_tipo_apoyo = request.form['nuevo-tipo-apoyo']
    nuevo_criterio_seleccion = request.form['nuevo-criterio-seleccion']
    nuevo_periodo_apoyo = request.form['nuevo-periodo-apoyo']
    nueva_fecha_asignada = request.form['nueva-fecha-registro']
    fecha_obj = datetime.strptime(nueva_fecha_asignada, '%Y-%m-%d')
    
    if nuevo_nombre_escuela != None and nuevo_nombre_escuela != "":
        Escolares.nombre_escuela = nuevo_nombre_escuela
    if nuevo_cct != None and nuevo_cct != "":
        Escolares.cct = nuevo_cct
    if nueva_localidad_escuela != None and nueva_localidad_escuela != "":
        Escolares.localidad_escuela = nueva_localidad_escuela
    if nuevo_nombre_beneficiario != None and nuevo_nombre_beneficiario != "":
        Escolares.nombre_beneficiario = nuevo_nombre_beneficiario
    if  nueva_fecha_otj != None and  nueva_fecha_otj != "":
        Escolares.fecha_nacimiento =  nueva_fecha_otj
    if nuevo_genero_beneficiario != None and nuevo_genero_beneficiario != "":
        Escolares.genero = nuevo_genero_beneficiario
    if nuevo_estado_civil != None and nuevo_estado_civil != "":
        Escolares.estado_civil = nuevo_estado_civil
    if nueva_escolaridad_beneficiario != None and nueva_escolaridad_beneficiario != "":
        Escolares.escolaridad_beneficiario = nueva_escolaridad_beneficiario
    if nuevo_ine_beneficiario != None and nuevo_ine_beneficiario != "":
        Escolares.ine_beneficiario = nuevo_ine_beneficiario
    if nueva_curp_beneficiario != None and nueva_curp_beneficiario != "":
        Escolares.curp_beneficiario = nueva_curp_beneficiario
    if nuevo_telefono != None and nuevo_telefono != "":
        Escolares.numero_telefono = nuevo_telefono
    if nuevo_correo != None and nuevo_correo != "":
        Escolares.correo_beneficiario = nuevo_correo
    if nuevo_municipio != None and nuevo_municipio != "":
        Escolares.municipio_beneficiario = nuevo_municipio
    if nueva_localidad != None and nueva_localidad != "":
        Escolares.localidad_beneficiario = nueva_localidad
    if nueva_colonia != None and nueva_colonia != "":
        Escolares.colonia_beneficiario = nueva_colonia
    if nuevo_codigo_postal != None and nuevo_codigo_postal != "":
        Escolares.codigo_postal = nuevo_codigo_postal
    if nuevo_numero_exterior != None and nuevo_numero_exterior != "":
        Escolares.numero_exterior = nuevo_numero_exterior
    if nuevo_numero_interior != None and nuevo_numero_interior != "":
        Escolares.numero_interior = nuevo_numero_interior
    if nueva_calle != None and nueva_calle != "":
        Escolares.calle = nueva_calle
    if nueva_referencia_domicilio != None and nueva_referencia_domicilio != "":
        Escolares.referencias_domicilio = nueva_referencia_domicilio
    if nuevo_tipo_apoyo != None and nuevo_tipo_apoyo != "":
        Escolares.tipo_apoyo = nuevo_tipo_apoyo
    if nuevo_criterio_seleccion != None and nuevo_criterio_seleccion != "":
        Escolares.criterio_seleccion = nuevo_criterio_seleccion
    if nuevo_periodo_apoyo != None and nuevo_periodo_apoyo != "":
        Escolares.periodo_apoyo = nuevo_periodo_apoyo
    if fecha_obj != None and fecha_obj != "":
        Escolares.fecha_registro = fecha_obj
    
    db_session.add(Escolares)
    db_session.commit()

    return redirect(url_for('escolares'))

@app.get('/huertos_familiares/')
def familiares():
    familiares = db_session.query(models.Familiares).all()
    user_type = session.get('user_type')
    
    # Obtener el parámetro 'per_page' de la solicitud GET
    per_page = request.args.get('per_page', type=int, default=5)
    page = request.args.get('page', type=int, default=1)
    
    total_records = db_session.query(func.count(models.Familiares.id_beneficiario)).scalar()  # Obtener el número total de registros

    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    total_pages = total_records // per_page
    if total_records % per_page != 0:
        total_pages += 1

    if page < 1:
        page = 1
    elif page > total_pages:
        page = total_pages

    if user_type == 'admin':
        familiares_paginados = db_session.query(models.Familiares).slice(start_index, end_index)
        return render_template('modulo-huerto-familiar.html', Familiares=familiares_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)

    modulos_accesibles = session.get('modulos', [])
    if 'Huertos familiares' in modulos_accesibles:
        familiares_paginados = db_session.query(models.Familiares).slice(start_index, end_index)
        return render_template('modulo-huerto-familiar.html', Familiares=familiares_paginados, page=page, total_pages=total_pages, per_page=per_page, total_records=total_records)
    else:
        return redirect(url_for('login_form'))
    
#formulario huertos familiares
@app.get('/huertos/familiares/form')
def familiares_form():
    familiares = db_session.query(models.Familiares).all()
    user_type = session.get('user_type')

    if user_type == 'admin':
        return render_template('formulario-huertos-familiares.html', Familiares=familiares)

    modulos_accesibles = session.get('modulos', [])
    if 'Huertos escolares' in modulos_accesibles:
        return render_template('formulario-huertos-familiares.html', Familiares=familiares)
    else:
        return redirect(url_for('login_form'))
    
#post
@app.post('/huertos/familiares/post')
def familiares_post():
    page = request.args.get('page', type=int)
    nombre_beneficiario = request.form['nombre_beneficiario']
    fecha_str = request.form['fecha_nacimiento']
    fecha_otj = datetime.strptime(fecha_str, '%Y-%m-%d')
    genero = request.form['genero']
    estado_civil = request.form['estado_civil']
    escolaridad_beneficiario = request.form['escolaridad_beneficiario']
    ine_beneficiario = request.form['ine_beneficiario']
    curp_beneficiario = request.form['curp_beneficiario']
    numero_telefono = request.form['numero_telefono']
    correo_beneficiario = request.form['correo_beneficiario']
    municipio_beneficiario = request.form['municipio_beneficiario']
    localidad_beneficiario = request.form['localidad_beneficiario']
    colonia_beneficiario = request.form['colonia_beneficiario']
    codigo_postal = request.form['codigo_postal']
    numero_exterior = request.form['numero_exterior']
    numero_interior = request.form['numero_interior']
    calle = request.form['calle']
    referencias_domicilio = request.form['referencias_domicilio']
    tipo_apoyo = request.form['tipo_apoyo']
    criterio_seleccion = request.form['criterio_seleccion']
    periodo_apoyo = request.form['periodo_apoyo']
    fecha_asignada = request.form['fecha_registro']
    fecha_obj = datetime.strptime(fecha_asignada, '%Y-%m-%d')
    
    registro_nuevo =models.Familiares(
        nombre_beneficiario = nombre_beneficiario,
        fecha_nacimiento = fecha_otj,
        genero = genero,
        estado_civil = estado_civil,
        escolaridad_beneficiario = escolaridad_beneficiario,
        ine_beneficiario = ine_beneficiario,
        curp_beneficiario = curp_beneficiario,
        numero_telefono =numero_telefono,
        correo_beneficiario = correo_beneficiario,
        municipio_beneficiario = municipio_beneficiario,
        localidad_beneficiario = localidad_beneficiario,
        colonia_beneficiario = colonia_beneficiario,
        codigo_postal = codigo_postal,
        numero_exterior =  numero_exterior,
        numero_interior = numero_interior,
        calle = calle,
        referencias_domicilio = referencias_domicilio,
        tipo_apoyo = tipo_apoyo,
        criterio_seleccion = criterio_seleccion,
        periodo_apoyo = periodo_apoyo,
        fecha_registro = fecha_obj
        
    )
    
    db_session.add(registro_nuevo)
    db_session.commit()
    return redirect(url_for('familiares', page=page))

#Eliminar
@app.get('/huertos/familiares/<id>/delete')
def familiares_delete(id):
    familiares = db_session.query(models.Familiares).get(id)
    db_session.delete(familiares)
    db_session.commit()
    return redirect(url_for('familiares'))

#actualización
@app.post('/huertos/familiares/<id>/update')
def familiares_update(id):
    Familiares = db_session.query(models.Familiares).get(id)
    nuevo_nombre_beneficiario = request.form['nuevo-nombre-beneficiario']
    nueva_fecha_str = request.form['nueva-fecha-nacimiento']
    nueva_fecha_otj = datetime.strptime(nueva_fecha_str, '%Y-%m-%d')
    nuevo_genero_beneficiario = request.form['nuevo-genero-beneficiario']
    nuevo_estado_civil = request.form['nuevo-estado-civil']
    nueva_escolaridad_beneficiario = request.form['nueva-escolaridad-beneficiario']
    nuevo_ine_beneficiario = request.form['nuevo-ine-beneficiario']
    nueva_curp_beneficiario = request.form['nueva-curp-beneficiario']
    nuevo_telefono = request.form['nuevo-telefono']
    nuevo_correo = request.form['nuevo-correo']
    nuevo_municipio = request.form['nuevo-municipio']
    nueva_localidad = request.form['nueva-localidad']
    nueva_colonia = request.form['nueva-colonia']
    nuevo_codigo_postal = request.form['nuevo-codigo-postal']
    nuevo_numero_exterior = request.form['nuevo-numero-exterior']
    nuevo_numero_interior = request.form['nuevo-numero-interior']
    nueva_calle = request.form['nueva-calle']
    nueva_referencia_domicilio = request.form['nueva-referencia-domicilio']
    nuevo_tipo_apoyo = request.form['nuevo-tipo-apoyo']
    nuevo_criterio_seleccion = request.form['nuevo-criterio-seleccion']
    nuevo_periodo_apoyo = request.form['nuevo-periodo-apoyo']
    nueva_fecha_asignada = request.form['nueva-fecha-registro']
    fecha_obj = datetime.strptime(nueva_fecha_asignada, '%Y-%m-%d')
    
    if nuevo_nombre_beneficiario != None and nuevo_nombre_beneficiario != "":
        Familiares.nombre_beneficiario = nuevo_nombre_beneficiario
    if  nueva_fecha_otj != None and  nueva_fecha_otj != "":
        Familiares.fecha_nacimiento =  nueva_fecha_otj
    if nuevo_genero_beneficiario != None and nuevo_genero_beneficiario != "":
        Familiares.genero = nuevo_genero_beneficiario
    if nuevo_estado_civil != None and nuevo_estado_civil != "":
        Familiares.estado_civil = nuevo_estado_civil
    if nueva_escolaridad_beneficiario != None and nueva_escolaridad_beneficiario != "":
        Familiares.escolaridad_beneficiario = nueva_escolaridad_beneficiario
    if nuevo_ine_beneficiario != None and nuevo_ine_beneficiario != "":
        Familiares.ine_beneficiario = nuevo_ine_beneficiario
    if nueva_curp_beneficiario != None and nueva_curp_beneficiario != "":
        Familiares.curp_beneficiario = nueva_curp_beneficiario
    if nuevo_telefono != None and nuevo_telefono != "":
        Familiares.numero_telefono = nuevo_telefono
    if nuevo_correo != None and nuevo_correo != "":
        Familiares.correo_beneficiario = nuevo_correo
    if nuevo_municipio != None and nuevo_municipio != "":
        Familiares.municipio_beneficiario = nuevo_municipio
    if nueva_localidad != None and nueva_localidad != "":
        Familiares.localidad_beneficiario = nueva_localidad
    if nueva_colonia != None and nueva_colonia != "":
        Familiares.colonia_beneficiario = nueva_colonia
    if nuevo_codigo_postal != None and nuevo_codigo_postal != "":
        Familiares.codigo_postal = nuevo_codigo_postal
    if nuevo_numero_exterior != None and nuevo_numero_exterior != "":
        Familiares.numero_exterior = nuevo_numero_exterior
    if nuevo_numero_interior != None and nuevo_numero_interior != "":
        Familiares.numero_interior = nuevo_numero_interior
    if nueva_calle != None and nueva_calle != "":
        Familiares.calle = nueva_calle
    if nueva_referencia_domicilio != None and nueva_referencia_domicilio != "":
        Familiares.referencias_domicilio = nueva_referencia_domicilio
    if nuevo_tipo_apoyo != None and nuevo_tipo_apoyo != "":
        Familiares.tipo_apoyo = nuevo_tipo_apoyo
    if nuevo_criterio_seleccion != None and nuevo_criterio_seleccion != "":
        Familiares.criterio_seleccion = nuevo_criterio_seleccion
    if nuevo_periodo_apoyo != None and nuevo_periodo_apoyo != "":
        Familiares.periodo_apoyo = nuevo_periodo_apoyo
    if fecha_obj != None and fecha_obj != "":
        Familiares.fecha_registro = fecha_obj
    
    db_session.add(Familiares)
    db_session.commit()

    return redirect(url_for('familiares'))

#Reporte en excel desayunos frios
@app.route('/generar_reporte_excel', methods=['GET', 'POST'])
def generar_reporte_excel():
    datos_frios = db_session.query(models.Frio).all()

    # Crear un nuevo libro de trabajo de Excel y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
 # Fusionar celdas para el título en la columna F
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=15) 

    # Ajustar el espacio en blanco entre el título y el encabezado
    for _ in range(2):
        ws.append([])

    # Dar estilo al título
    title = ws.cell(row=2, column=6)  # Empieza en la fila 3 y columna 6 (F)
    title.value = "Reporte Desayuno escolar frio"
    title.font = Font(bold=True, size=16)  
    title.alignment = Alignment(horizontal="center", vertical="center")

     # Aumentar el espacio entre celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

        # Agregar encabezados
    encabezados = ['Escuela', 'CCT', 'Localidad', 'Beneficiario', 'Escolaridad', 'Talla', 'CURP', 'Peso', 'Referencias domicilio', 'Padre o tutot', 'Escolaridad', 'INE', 'CURP', 'Teléfono', 'Estado Civil']
    for col, header in enumerate(encabezados, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF") 
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Fondo negro

    # Agregar los datos de beneficiarios a la hoja de cálculo
    for row, frio in enumerate(datos_frios, 5):
        fila = [
            frio.nombre_escuela,
            frio.cct,
            frio.localidad_escuela,
            frio.nombre_beneficiario,
            frio.escolaridad_beneficiario,
            frio.talla_beneficiario,
            frio.curp_beneficiario,
            frio.peso_beneficiario,
            frio.referencias_domicilio,
            frio.nombre_padre,
            frio.escolaridad_padre,
            frio.ine_padre,
            frio.curp_padre,
            frio.telefono_padre,
            frio.estado_civil
        ]
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=row, column=col, value=valor)

    # Aumentar el alto de las filas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'inline; filename=desayunos_frios.xlsx'

    return response

#Reporte en excel desayunos calientes
@app.route('/reporte/desayunos/calientes', methods=['GET', 'POST'])
def comunitario_excel():
    # Recupera los datos de labase de datos
    desayunos_calientes = db_session.query(models.Comunitario).all()
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
 
    # Fusionar celdas para el título en la columna F
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=15) 

    for _ in range(2):
        ws.append([])

    # Dar estilo al título
    title = ws.cell(row=2, column=6)
    title.value = "Reporte Desayuno Escolar Comunitario"
    title.font = Font(bold=True, size=16)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Aumentar el espacio entre celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Agregar encabezados
    encabezados = ['Escuela', 'CCT', 'Localidad', 'Beneficiario', 'Escolaridad', 'Talla', 'CURP', 'Peso', 'Referencias domicilio', 'Padre o tutot', 'Escolaridad', 'INE', 'CURP', 'Teléfono', 'Estado Civil']
    for col, header in enumerate(encabezados, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # Agregar los datos de beneficiarios a la hoja de cálculo
    for row, comunitario in enumerate(desayunos_calientes, 5):
        fila = [
            comunitario.nombre_escuela,
            comunitario.cct,
            comunitario.localidad_escuela,
            comunitario.nombre_beneficiario,
            comunitario.escolaridad_beneficiario,
            comunitario.talla_beneficiario,
            comunitario.curp_beneficiario,
            comunitario.peso_beneficiario,
            comunitario.referencias_domicilio,
            comunitario.nombre_padre,
            comunitario.escolaridad_padre,
            comunitario.ine_padre,
            comunitario.curp_padre,
            comunitario.telefono_padre,
            comunitario.estado_civil
        ]
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=row, column=col, value=valor)

    # Aumentar el alto de las filas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Configurar los encabezados de respuesta para abrir el archivo en el navegador
    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'inline; filename=desayunos_comunitarios.xlsx'

    return response

#Reporte en excel huertos escolares
@app.route('/reporte/huertos/escolares', methods=['GET', 'POST'])
def escolares_excel():
    # Recupera los datos de la base de datos
    huertos_escolares = db_session.query(models.Escolares).all()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['T'].width = 20
    ws.column_dimensions['U'].width = 20
    ws.column_dimensions['V'].width = 20
    ws.column_dimensions['W'].width = 20
 
    # Fusionar celdas
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=23)

    for _ in range(2):
        ws.append([])

    # Dar estilo al título
    title = ws.cell(row=2, column=6) 
    title.value = "Reporte huertos escolares"
    title.font = Font(bold=True, size=16)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Aumentar el espacio entre celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Agregar encabezados
    encabezados = ['Escuela', 'CCT', 'Localidad', 'Beneficiario', 'Fecha de nacimiento', 'Genero', 'Estado civil', 'Escolaridad', 'INE', 'CURP', 'Telefono', 'Correo', 'Municipio', 'Localidad', 'Colonia', 'NE', 'NI', 'Calle', 'Referencias domicilio', 'Tipo de apoyo', 'Criterio de selección', 'Periodo', 'Fecha de registro']
    for col, header in enumerate(encabezados, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # Agregar los datos de beneficiarios a la hoja de cálculo
    for row, escolares in enumerate(huertos_escolares, 5):
        fila = [
            escolares.nombre_escuela,
            escolares.cct,
            escolares.localidad_escuela,
            escolares.nombre_beneficiario,
            escolares.fecha_nacimiento,
            escolares.genero,
            escolares.estado_civil,
            escolares.escolaridad_beneficiario,
            escolares.ine_beneficiario,
            escolares.curp_beneficiario,
            escolares.numero_telefono,
            escolares.correo_beneficiario,
            escolares.municipio_beneficiario,
            escolares.localidad_beneficiario,
            escolares.colonia_beneficiario,
            escolares.numero_exterior,
            escolares.numero_interior,
            escolares.calle,
            escolares.referencias_domicilio,
            escolares.tipo_apoyo,
            escolares.criterio_seleccion,
            escolares.periodo_apoyo,
            escolares.fecha_registro
        ]
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=row, column=col, value=valor)

    # Aumentar el alto de las filas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'inline; filename=huertos_escolares.xlsx'

    return response

#Reporte en excel huertos familiares
@app.route('/reporte/huertos/familiares', methods=['GET', 'POST'])
def familiares_excel():
    # Recupera los datos de la base de datos
    huertos_familiares = db_session.query(models.Familiares).all()

    # Crear un nuevo libro de trabajo de Excel y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['T'].width = 20
 
    # Fusionar celdas
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=20)

    for _ in range(2):
        ws.append([])

    # Dar estilo al título
    title = ws.cell(row=2, column=6)
    title.value = "Reporte huertos familiares"
    title.font = Font(bold=True, size=16)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Aumentar el espacio entre celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Agregar encabezados
    encabezados = ['Beneficiario', 'Fecha de nacimiento', 'Genero', 'Estado civil', 'Escolaridad', 'INE', 'CURP', 'Telefono', 'Correo', 'Municipio', 'Localidad', 'Colonia', 'NE', 'NI', 'Calle', 'Referencias domicilio', 'Tipo de apoyo', 'Criterio de selección', 'Periodo', 'Fecha de registro']
    for col, header in enumerate(encabezados, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF") 
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # Agregar los datos de beneficiarios a la hoja de cálculo
    for row, familiares in enumerate(huertos_familiares, 5):
        fila = [
            familiares.nombre_beneficiario,
            familiares.fecha_nacimiento,
            familiares.genero,
            familiares.estado_civil,
            familiares.escolaridad_beneficiario,
            familiares.ine_beneficiario,
            familiares.curp_beneficiario,
            familiares.numero_telefono,
            familiares.correo_beneficiario,
            familiares.municipio_beneficiario,
            familiares.localidad_beneficiario,
            familiares.colonia_beneficiario,
            familiares.numero_exterior,
            familiares.numero_interior,
            familiares.calle,
            familiares.referencias_domicilio,
            familiares.tipo_apoyo,
            familiares.criterio_seleccion,
            familiares.periodo_apoyo,
            familiares.fecha_registro
        ]
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=row, column=col, value=valor)

    # Aumentar el alto de las filas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'inline; filename=huertos_familiares.xlsx'

    return response

@app.post('/buscar/beneficiario/frio')
def buscar_frios():
    
    busqueda = request.form['buscar']
    if busqueda != None or busqueda != '':
        resultados = buscar_beneficiario(busqueda)

    alert = ''

    if len(resultados) > 0:
        print("Resultados:")
    else:
        alert = f"No se encontro resultados de :{busqueda}"

    return render_template('modulo-desayuno-frio.html', frio = resultados, alerta = alert)

def buscar_beneficiario(busqueda):
    frios = db_session.query(models.Frio).filter(
        models.Frio.nombre_beneficiario.like(f'%{busqueda}%')
    ).all()
    return frios

app.run("0.0.0.0",8080,debug=True)